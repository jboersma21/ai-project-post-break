"""
Project Post Break

Team 6

Colin Moody, Ohad Beck, Charlie MacVicar, Jake Boersma

"""

import sys
import heapq
from openpyxl import load_workbook

import data_import
from world_objects import *
from config import *
import pprint


# Implement state manager to traverse through of current state, future states, and previous states
class WorldStateManager(object):

    def __init__(self, depth_bound, frontier_max_size, initial_resources, initial_countries, num_output_schedules):
        self.cur_state = World(d_bound=depth_bound, weight_dict=initial_resources, country_dict=initial_countries)
        self.frontier_queue = list()  # priority queue that store states
        self.prev_states = list()  # stack of explored states
        self.cur_depth = 0
        self.depth_bound = depth_bound
        self.frontier_max_size = frontier_max_size
        self.inter_prob_success = 0
        self.init_output_schedules = num_output_schedules
        self.output_schedules_left = num_output_schedules
        self.completed_sched_eu = []


    # unfinished depth-bound search algorithm
    def execute_search(self):
        self.cur_state.prev_op.append("None: State 0")
        heapq.heappush(self.frontier_queue, (self.cur_state.exp_utility * -1, self.cur_state))

        while self.output_schedules_left > 0:
            # if self.cur_depth == self.depth_bound:
            #     self.cur_depth = 1
            #     prev_eu = 0
            #     #self.cur_state = self.pop_unexplored_state(1)
            #     self.print_cur_state_info()
            #     print("Marginal Utility: ", (self.get_cur_eu() - prev_eu))
            #     prev_eu = self.get_cur_eu()

            self.go_to_next_state()
            #self.print_cur_state_info()
            #print("Marginal Utility: ", (self.get_cur_eu() - prev_eu))
            #prev_eu = self.get_cur_eu()
        # to-do: add depth-bounded logic

    def go_to_next_state(self):
        successors_queue = list()
        self.cur_depth += 1          # increment depth before updating values?
        self.cur_state = self.pop_future_state()
        for world in generate_successors(self.cur_state):
            world.search_depth += 1
            for country in world.countries:                     # update all measures for each successor before adding future state
                world.countries[country].update_discount_reward(world.search_depth)
                world.countries[country].update_c_prob_success()
                world.prob_success = world.prob_success * world.countries[country].c_prob_success
            world.update_exp_utility()
            self.inter_prob_success = world.prob_success
            world.prob_success = 1  # update back to 1 after calculating EU

            # self.add_future_state(world_state=world)
            heapq.heappush(successors_queue, (world.exp_utility * -1, world))
        self.prev_states.append(self.get_cur_eu())
        while len(self.frontier_queue) < self.frontier_max_size:
            next_successor = heapq.heappop(successors_queue)[1]
            next_successor.prev_eu.append(self.get_cur_eu())
            cur_eu = next_successor.exp_utility
            if next_successor.search_depth == self.depth_bound:
                if cur_eu in self.completed_sched_eu:
                    return
                else:
                    next_successor.prev_eu.append(cur_eu)
                    self.completed_sched_eu.append(cur_eu)
                    self.output_schedules_left -= 1
                    print("Schedule #: ", self.init_output_schedules - self.output_schedules_left)
                    for i in range(0, len(next_successor.prev_eu)):
                        print("\nSearch Depth: ", i)
                        print('\t', next_successor.prev_op[i])
                        print('\t', next_successor.prev_eu[i])
                    print("\n\n")
                    return
            else:

                heapq.heappush(self.frontier_queue, (next_successor.exp_utility * -1, next_successor))

    #
    # def add_future_state(self, world_state):
    #     # maintain prio queue max size
    #     #      --> if max size reached, add state if better EU than worst currently in queue
    #     if len(self.frontier_queue) == self.frontier_max_size:
    #         raise ValueError('Error: frontier max size exceeded')
    #     elif len(self.frontier_queue) == self.frontier_max_size:
    #         if (world_state.exp_utility * -1) < heapq.nlargest(1, self.frontier_queue)[0][0]:
    #             self.frontier_queue.remove(heapq.nlargest(1, self.frontier_queue)[0])
    #             heapq.heappush(self.frontier_queue, (world_state.exp_utility * -1, world_state))
    #     else:
    #         heapq.heappush(self.frontier_queue, (world_state.exp_utility * -1, world_state))


    def pop_future_state(self):
        return heapq.heappop(self.frontier_queue)[1]

    def print_cur_state_info(self):
        if self.cur_depth > 0:
            print('\t-> Operator: {}'.format(self.cur_state.prev_op))
        print('State {}:\t{}'.format(self.cur_depth, self.get_cur_eu()))
        print('Prob World: {}'.format(self.inter_prob_success))
        if self.cur_depth > 0:
            print('Prob Self: ', self.cur_state.self_country.c_prob_success)
            print('DR: ', self.cur_state.self_country.discount_reward)

    def get_cur_eu(self):
        return self.cur_state.exp_utility

def generate_successors(current_state):
    successors = list()
    # Add every transformation for every country
    for country in current_state.countries.keys():
        for operator in configuration["definitions"]:
            if operator != 'transfer':                          # FIX LATER (MAYBE DONT READ IN TRANSFERS)
                tmp_world = current_state.get_deep_copy()
                bins = 1
                while tmp_world.countries[country].transform(transformation=operator, bins=bins):
                    ins = [i * bins for i in configuration['definitions'][operator]["in"].values()]
                    outs = [i * bins for i in configuration['definitions'][operator]["out"].values()]
                    tmp_world.update_prev_op('{} (in={} out={}) (bins={}) country={}'.format(operator, ins, outs, bins, country))
                    successors.append(tmp_world)
                    bins += 5
                    tmp_world = current_state.get_deep_copy()

    # Add every transfer for every pair of countries (both ways)
    for exporter in current_state.countries.keys():
        for destination in current_state.countries.keys():
            if exporter != destination:
                for resource in configuration["resources"]:
                    tmp_world = current_state.get_deep_copy()
                    bins = 1
                    while tmp_world.transfer(exporter=exporter, destination=destination, resource=resource, bins=bins):
                        tmp_world.update_prev_op('{} (from={} to={} resource={} amount={})'
                                              ''.format('transfer', exporter, destination, resource, bins))
                        successors.append(tmp_world)
                        bins += 5
                        tmp_world = current_state.get_deep_copy()

    return successors

def output_successors_to_excel(file_name, successors):
    wb = load_workbook(file_name)
    sheet_name = 'Successors (Test Results)'
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        wb.remove(sheet)
    ws = wb.create_sheet(sheet_name)
    cur_row = 1
    cur_col = 1

    for idx, state in enumerate(successors):
        if idx > 9:  # only print first 10 successors
            break
        ws.cell(row=cur_row, column=cur_col).value = 'Successor'
        cur_col += 1
        ws.cell(row=cur_row, column=cur_col).value = idx + 1
        cur_col += 2
        ws.cell(row=cur_row, column=cur_col).value = 'Big-U'
        cur_col += 1
        ws.cell(row=cur_row, column=cur_col).value = state.get_big_u()
        cur_col += 2
        ws.cell(row=cur_row, column=cur_col).value = 'Prev Op'
        cur_col += 1
        ws.cell(row=cur_row, column=cur_col).value = state.prev_op

        cur_row += 1
        cur_col = 1
        ws.cell(row=cur_row, column=cur_col).value = 'Country'
        cur_col += 1
        for r in configuration['resources']:
            ws.cell(row=cur_row, column=cur_col).value = r
            cur_col += 1

        cur_row += 1
        cur_col = 1
        for name, country in state.countries.items():
            ws.cell(row=cur_row, column=cur_col).value = name
            cur_col += 1
            for r in configuration['resources']:
                ws.cell(row=cur_row, column=cur_col).value = country.resources[r]
                cur_col += 1
            cur_col = 1
            cur_row += 1

        cur_row += 1

    print('{} Test Complete'.format(file_name))
    wb.save(file_name)



def run_successor_test(resources_filename, initial_state_filename, operator_def_filename, output_schedule_filename):
    data_import.read_operator_def_config(file_name=operator_def_filename)
    my_state_manager = WorldStateManager(depth_bound=3,
                                         frontier_max_size=100,
                                         initial_resources=data_import.create_resource_dict(file_name=resources_filename),
                                         initial_countries=data_import.create_country_dict(file_name=initial_state_filename),
                                         num_output_schedules=25)
    #output_successors_to_excel(file_name=output_schedule_filename, successors=generate_successors(my_state_manager.cur_state))

    print('\nExample Search on {}:'.format(resources_filename))
    my_state_manager.execute_search()
    print('\n')


def main(argv):
    # for name in ['Test1', 'Test2', 'Test3', 'Test4']:
    #     run_successor_test(file_name='data/Pre_Break/Initial-World-{}.xlsx'.format(name))

    for name in ['1']:
        run_successor_test(resources_filename='data/resources_{}.xlsx'.format(name),
                           initial_state_filename='data/initial_state_{}.xlsx'.format(name),
                           operator_def_filename='data/operator_def_{}.xlsx'.format(name),
                           output_schedule_filename='data/output_{}.xlsx'.format(name))


if __name__ == "__main__":
    main(sys.argv)
