"""
Project Post Break

Team 6

Colin Moody, Ohad Beck, Charlie MacVicar, Jake Boersma

"""
import string
from openpyxl import load_workbook
from config import *
import re
from pprint import *

DATA_FILE = 'data/Pre_Break/Initial-World.xlsx'

RESOURCE_DATA_FILE = 'data/resources.xlsx'
INITIAL_STATE_DATA_FILE = 'data/initial_state.xlsx'
OPERATOR_DEF_DATA_FILE = 'data/operator_def.xlsx'


def create_resource_dict(file_name=RESOURCE_DATA_FILE):
    """
    Creates a dictionary containing resource and weight info.
    Keys: name of resource
    Values: weight of corresponding resource
    """
    wb = load_workbook(file_name)
    resources_sheet = wb['Resources']
    resource_dict = {}
    for row in range(2, resources_sheet.max_row + 1):
        key = get_val(resources_sheet, 'A', row)
        value = get_val(resources_sheet, 'B', row)
        if key is not None:
            configuration["resources"].append(key)             # adding resources to config file
            resource_dict[key] = value
    return resource_dict


def create_country_dict(file_name=INITIAL_STATE_DATA_FILE):
    """
    Creates a dictionary containing country info.
    Keys: name of country
    Values: dictionary where key = resource_name and value = resource_amount
    """
    wb = load_workbook(file_name)
    country_sheet = wb['Countries']
    country_dict = {}
    for row in range(2, country_sheet.max_row + 1):
        key = get_val(country_sheet, 'A', row)
        self_val = get_val(country_sheet, 'B', row)
        resrc_dict = {}
        for col in range(3, country_sheet.max_column + 1):
            resrc_key = get_val(country_sheet, col_letter(col), 1)
            resrc_value = get_val(country_sheet, col_letter(col), row)
            resrc_dict[resrc_key] = resrc_value
        country_dict[key] = resrc_dict
        if self_val == 'Yes':
            country_dict['Self'] = key
    return country_dict


"""
    use operator_def_1.xlsx
"""
def read_operator_def_config(file_name=OPERATOR_DEF_DATA_FILE):
    wb = load_workbook(file_name)
    op_sheet = wb['Operators']

    for row in range(2, op_sheet.max_row + 1):
        op_type = get_val(op_sheet, 'A', row)       # Operation type (transfer or transform)
        op_name = get_val(op_sheet, 'B', row)       # Name of the operation (alloys_transform)
        op_str = get_val(op_sheet, 'C', row)        # Actual operator definition

        if op_str is not None:
            configuration["operations"].append(op_name)
            split_lst = re.split(r'[()\s]\s*', str(op_str))
            loop = True
            while loop:
                try:
                    split_lst.remove('')
                except ValueError:
                    loop = False
        else:
            return

        if op_type == 'TRANSFER':
            configuration["definitions"].update({op_name: {"from": split_lst[1], "to": split_lst[2],
                                                        "resrc": str(split_lst[3]), "amt": split_lst[4]}})

        else:
            configuration["definitions"].update({op_name: {"in": {}, "out": {}}})

            input_idx = 0
            output_idx = 0
            counter = 0
            for token in split_lst:
                if token == 'INPUTS':
                    input_idx = counter
                if token == 'OUTPUTS':
                    output_idx = counter
                counter += 1

            for i in range(input_idx + 1, output_idx, 2):
                configuration["definitions"][op_name]["in"].update({split_lst[i]: int(split_lst[i + 1])})

            for j in range(output_idx + 1, len(split_lst), 2):
                configuration["definitions"][op_name]["out"].update({split_lst[j]: int(split_lst[j + 1])})

def col_letter(col_num):
    letter_dict = dict(enumerate(string.ascii_uppercase, 1))
    return letter_dict[col_num]


def get_val(sheet, col, row):
    """Returns the value of a cell."""
    return sheet[col + str(row)].value


def print_resource_dict(dic):
    """Prints a resource dictionary's key and values line by line."""
    for key in dic:
        print(key, dic[key])


def print_country_dict(dic):
    """Prints a country dictionary's key and values line by line."""
    for key in dic:
        dic2 = dic[key]
        print(key)
        for key2 in dic2:
            print('\t' + key2, dic2[key2])


if __name__ == '__main__':
    pprint(create_resource_dict('data/resources_1.xlsx'))
    pprint(create_country_dict('data/initial_state_1.xlsx'))
    read_operator_def_config('data/operator_def_1.xlsx')
    pprint(configuration)



